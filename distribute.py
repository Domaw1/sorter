import csv
import shutil
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from utils.file_parser import (
    extract_project_code,
    match_pattern,
    extract_section_code,
    extract_subfolder_code,
    extract_revision,
    extract_extension_file, get_project_folder_name,
)

from utils.crc import crc32_of_file
from utils.logger import setup_logger

import os
import config

def build_target_path(category, section, subfolder, revision, file_type):
    base = os.path.join(config.TARGET_DIR, category, section, subfolder)

    if revision:
        base = os.path.join(base, f"Рев. {revision}", file_type)
    else:
        base = os.path.join(base, file_type)

    return base


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

# def write_report_xlsx(rows, filename):
#     os.makedirs("reports", exist_ok=True)
#     report_path = os.path.join("reports", filename)
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчёт"
#
#     headers = [
#         "Имя файла",
#         "Исходный путь",
#         "Конечный путь",
#         "Категория",
#         "Раздел",
#         "Подпапка",
#         "Ревизия",
#         "Тип файла",
#         "CRC",
#         "Статус",
#         "Пропущенный файл"
#     ]
#
#     # Записываем заголовки
#     for col, header in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col, value=header)
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center")
#
#     # Записываем строки
#     for row_idx, row in enumerate(rows, start=2):
#         for col_idx, value in enumerate(row, start=1):
#             ws.cell(row=row_idx, column=col_idx, value=value)
#
#     # Фиксированная ширина столбцов
#     column_widths = {
#         1: 70,   # Имя файла
#         2: 135,   # Исходный путь
#         3: 155,   # Конечный путь
#         4: 15,   # Категория
#         5: 10,   # Раздел
#         6: 40,   # Подпапка
#         7: 10,   # Ревизия
#         8: 15,   # Тип файла
#         9: 15,   # CRC
#         10: 50,  # Статус
#         11: 40 # Пропущенный файл
#     }
#
#     for col_idx, width in column_widths.items():
#         ws.column_dimensions[get_column_letter(col_idx)].width = width
#
#     wb.save(report_path)
#     return report_path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

def write_report_xlsx(processed_rows, skipped_rows, filename):

    os.makedirs("reports", exist_ok=True)
    report_path = os.path.join("reports", filename)

    wb = Workbook()

    # === Лист 1: Обработанные ===
    # === Лист 1: Обработанные ===
    ws_main = wb.active
    ws_main.title = "Обработанные файлы"

    headers_main = [
        "Имя файла", "Исходный путь", "Конечный путь", "Категория",
        "Раздел", "Подпапка","Исходная ревизия", "Исправленная ревизия", "Тип файла", "CRC", "Статус"
    ]

    for col, header in enumerate(headers_main, start=1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # === Лист 2: Пропущенные ===
    ws_skipped = wb.create_sheet(title="Пропущенные файлы")

    headers_skipped = ["Имя файла", "Исходный путь", "Причина пропуска"]
    for col, header in enumerate(headers_skipped, start=1):
        cell = ws_skipped.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # === Заполнение
    row_main = 2
    row_skipped = 2

    # for row in rows:
    #     status = row[9]
    #     if status == "Не было скопировано" or "не соответствует шаблону" in status.lower():
    #         # Пропущенные
    #         ws_skipped.cell(row=row_skipped, column=1, value=row[0])  # Имя файла
    #         ws_skipped.cell(row=row_skipped, column=2, value=row[1])  # Исходный путь
    #         ws_skipped.cell(row=row_skipped, column=3, value=status)  # Причина
    #         row_skipped += 1
    # for row in processed_rows:
    #     for col_idx in range(10):
    #         ws_main.cell(row=row_main, column=col_idx + 1, value=row[col_idx])
    #     row_main += 1
    #
    # for row in skipped_rows:
    #     for col_idx in range(3):
    #         ws_skipped.cell(row=row_skipped, column=col_idx + 1, value=row[col_idx])
    #     row_skipped += 1
    # else:
    #         # Обработанные
    #     for col_idx in range(10):
    #         ws_main.cell(row=row_main, column=col_idx + 1, value=row[col_idx])
    #     row_main += 1
    groups = {
        "Скопировано": [],
        "Дубликат (не копировался)": [],
        "Конфликт имён": []
    }

    # === Заполнение обработанных файлов с группировкой ===

    groups = {
        "Успешно скопировано": [],
        "В конечных папках уже есть данные файлы (дубликаты)": [],
        "Конфликты имён": []
    }

    for row in processed_rows:
        status = row[10]

        if "Скопировано" in status:
            groups["Успешно скопировано"].append(row)

        elif "Дубликат" in status:
            groups["В конечных папках уже есть данные файлы (дубликаты)"].append(row)

        elif "Конфликт" in status:
            groups["Конфликты имён"].append(row)

        else:
            groups.setdefault("Прочее", []).append(row)

    row_main = 2

    for group_name, group_rows in groups.items():
        # Заголовок группы
        cell = ws_main.cell(row=row_main, column=1, value=group_name + ":")
        cell.font = Font(bold=True)
        row_main += 1

        # Строки группы
        for row in group_rows:
            for col_idx in range(len(row)):
                ws_main.cell(row=row_main, column=col_idx + 1, value=row[col_idx])
            row_main += 1

        # Пустая строка между группами
        row_main += 1
        # === Заполнение листа "Пропущенные файлы" ===
        row_skipped = 2

        for row in skipped_rows:
            ws_skipped.cell(row=row_skipped, column=1, value=row[0])  # Имя файла
            ws_skipped.cell(row=row_skipped, column=2, value=row[1])  # Исходный путь
            ws_skipped.cell(row=row_skipped, column=3, value=row[2])  # Причина пропуска
            row_skipped += 1

    # === Ширина столбцов
    widths_main = [70, 135, 155, 15, 10, 40, 12, 12, 15, 15, 40]
    for i, w in enumerate(widths_main, start=1):
        ws_main.column_dimensions[get_column_letter(i)].width = w

    widths_skipped = [40, 60, 40]
    for i, w in enumerate(widths_skipped, start=1):
        ws_skipped.column_dimensions[get_column_letter(i)].width = w

    wb.save(report_path)
    return report_path


# def write_report_(rows, filename):
#     report_path = os.path.join("reports", filename)
#     os.makedirs("reports", exist_ok=True)
#
#     headers = [
#         "Имя файла",
#         "Исходный путь",
#         "Конечный путь",
#         "Категория",
#         "Раздел",
#         "Подпапка",
#         "Ревизия",
#         "Тип файла",
#         "CRC",
#         "Статус"
#     ]
#
#     with open(report_path, "w", encoding="utf-8-sig", newline="") as f:
#         writer = csv.writer(f, delimiter=";")
#         f.write('\ufeff')  # BOM уже есть, но можно оставить
#         f.write('# AUTOFIT_COLUMNS\n')
#         writer.writerow(headers)
#         writer.writerows(rows)
#
#     return report_path

def handle_duplicates(dst_file, src_file):
    src_crc = crc32_of_file(src_file)
    dst_crc = crc32_of_file(dst_file)

    if src_crc == dst_crc:
        return "duplicate", src_crc, dst_crc, dst_file, None

    base, ext = os.path.splitext(dst_file)
    new_dst = base + "_conflict" + ext
    logging.error("Name conflict: " + new_dst)
    return "conflict", src_crc, dst_crc, dst_file, new_dst

def process_file(folder_path, filename, category, report_rows, skipped_rows):
    src_file = os.path.join(folder_path, filename)

    # 1. Проверка шаблона
    # if not match_pattern(filename, config.NAME_PATTERN):
    #     logging.info(f"Пропуск: {filename} — не соответствует шаблону")
    #     return "skipped_pattern", filename, None

    if not match_pattern(filename, config.NAME_PATTERN):
        logging.info(f"Пропуск: {filename} — не соответствует шаблону")

        skipped_rows.append([
            filename,
            src_file,
            "Не соответствует шаблону имени"
        ])

        return "skipped_pattern", filename, None, None, None, False

    # 2. Разбор имени
    section = extract_section_code(filename)
    subfolder = extract_subfolder_code(filename)
    raw_revision, revision = extract_revision(filename)
    fixed = raw_revision and revision and raw_revision != revision
    ext = extract_extension_file(filename)

    if section is None:
        logging.error(f"Пропуск: {filename} — раздел не найден")
        return "error", filename, None, None, None, False

    if subfolder is None:
        logging.error(f"Пропуск: {filename} — подпапка не найдена")
        return "error", filename, None, None, None, False

    if ext is None:
        logging.error(f"Пропуск: {filename} — расширение не найдено")
        return "error", filename, None, None, None, False

    ext = ext.lower()
    if ext in config.BLACK_LIST_EXTENSIONS:
        logging.warning(f"Пропуск: {filename} — недопустимое расширение {ext}")
        return "warning", filename, None, None, None, False

    file_type = config.FILE_TYPE_MAP.get(ext, "ред.формат")
    if not file_type:
        logging.warning(f"Пропуск: {filename} — неизвестный тип файла")
        return "warning", filename, None, None, None, False

    # 3. Построение пути
    target_dir = build_target_path(category, section, subfolder, revision, file_type)
    ensure_dir(target_dir)
    dst_file = os.path.join(target_dir, filename)

    # 4. Проверка дубликатов
    status = "copied"
    src_crc = crc32_of_file(src_file)
    dst_crc = None

    if os.path.exists(dst_file):
        status, src_crc, dst_crc, old_dst, new_dst = handle_duplicates(dst_file, src_file)
        if status == "conflict":
            dst_file = new_dst
    else:
        old_dst = None
        new_dst = None

    # 5. Копирование
    if status in ("copied", "conflict"):
        ensure_dir(os.path.dirname(dst_file))
        shutil.copy2(src_file, dst_file)

    # 6. Логирование
    logging.info(
        f"{status.upper()} | {src_file} → {dst_file} | "
        f"Категория: {category} | Раздел: {section} | Подпапка: {subfolder} | Ревизия: {revision} | CRC: {src_crc}"
    )

    # 7. Отчёт
    report_rows.append([
        filename,
        src_file,
        dst_file,
        category,
        section,
        subfolder,
        raw_revision if raw_revision != revision else "",
        revision,
        file_type,
        src_crc,
        config.STATUS_LABELS.get(status, status),
        ""
    ])

    if status == "conflict":
        return status, old_dst, new_dst, raw_revision, revision, fixed
    else:
        return status, dst_file, None, raw_revision, revision, fixed

def collect_all_files(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        # если путь содержит ADRC — пропускаем
        if "ADRC-GPNG-GEP-RD" in dirpath:
            continue

        for filename in filenames:
            yield dirpath, filename

def main(progress_callback=None, stats_callback=None):
    setup_logger()
    report_rows = []
    skipped_rows = []

    stats = {
        "processed": 0,
        "copied": 0,
        "skipped_pattern": 0,
        "duplicates": 0,
        "conflicts": 0,
        "errors": 0,
        "report_path": "",
        "conflict_files": []
    }

    # Подсчёт общего количества файлов
    all_files = list(collect_all_files(config.SOURCE_DIR))
    total_files = len(all_files)

    # for folder in os.listdir(config.SOURCE_DIR):
    #     folder_path = os.path.join(config.SOURCE_DIR, folder)
    #     if not os.path.isdir(folder_path):
    #         continue
    #
    #     for f in os.listdir(folder_path):
    #         all_files.append((folder_path, f))

    processed = 0
    # valid_files = [
    #     (folder_path, filename)
    #     for folder_path, filename in all_files
    #     if extract_project_code(Path(folder_path).name) is not None
    # ]
    valid_files = [
        (folder_path, filename)
        for folder_path, filename in all_files
        if get_project_folder_name(Path(folder_path)) is not None
    ]

    total_files = len(valid_files)

    # for folder_path, filename in valid_files:
    #     # project_code = extract_project_code(os.path.basename(folder_path))
    #     project_code = extract_project_code(Path(folder_path).name)
    for folder_path, filename in valid_files:
        project_folder = get_project_folder_name(Path(folder_path))
        project_code = extract_project_code(project_folder)
        if project_code is None:
            continue

        category = config.PROJECT_MAP.get(project_code, config.DEFAULT_CATEGORY)
        # status, p1, p2 = process_file(folder_path, filename, category, report_rows)
        status, p1, p2, raw_revision, revision, fixed = process_file(folder_path, filename, category, report_rows, skipped_rows)

        stats["processed"] += 1

        if fixed:
            stats.setdefault("fixed_revisions", []).append({
                "file": filename,
                "raw": raw_revision,
                "normalized": revision
            })

        if status == "error":
            stats["errors"] += 1
        elif status == "skipped_pattern":
            stats["skipped_pattern"] += 1
        elif status == "duplicate":
            stats["duplicates"] += 1
        elif status == "conflict":
            stats["conflicts"] += 1
            stats["conflict_files"].append((p1, p2))
        elif status == "copied":
            stats["copied"] += 1

        processed += 1

        if progress_callback:
            progress_callback(processed, total_files, filename)

    # Отчёт
    report_name = f"report_{datetime.now():%Y-%m-%d}.xlsx"
    logging.info(f"Сохраняем отчёт: {report_name}")

    # report_path = write_report_(report_rows, report_name)
    report_path = write_report_xlsx(report_rows, skipped_rows, report_name)

    if stats_callback:
        stats["report_path"] = report_path
        stats_callback(stats)